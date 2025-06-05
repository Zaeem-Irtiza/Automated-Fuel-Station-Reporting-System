from openpyxl import load_workbook
from flask import Flask, render_template_string, request
import os

app = Flask(__name__)

# Load HTML template
with open('templates/index.html', 'r') as f:
    html_template = f.read()

def get_number(value):
    """Convert cell value to number safely"""
    if value is None:
        return 0
    try:
        return float(str(value).replace(',', ''))
    except (ValueError, TypeError):
        return 0

def process_sheet2(sheet1, sheet2, sheet3):
    """Process all Sheet2 calculations"""
    # For each row starting from row 2 in Sheet1 (Petrol)
    for current_row in range(2, sheet1.max_row):
        if current_row > 2:
            # Copy M-T columns to C-J columns
            for col in range(3, 10):  # Skip B (petrol)
                prev_final = sheet1.cell(row=current_row-1, column=col+10).value  # M-T columns
                if prev_final is not None:
                    sheet1.cell(row=current_row, column=col).value = prev_final  # C-J columns

    # Copy initial and final readings
    for col in range(2, 10):  # B through I
        sheet2.cell(row=8, column=col).value = sheet1.cell(row=2, column=col+1).value  # Initial
        sheet2.cell(row=7, column=col).value = sheet1.cell(row=2, column=col+11).value  # Final
        sheet2.cell(row=9, column=col).value = get_number(sheet2.cell(row=7, column=col).value) - get_number(sheet2.cell(row=8, column=col).value)

    # Calculate totals
    quantity_sold = get_number(sheet2['B9'].value) + get_number(sheet2['C9'].value)
    all_nozzles = sum(get_number(sheet2.cell(row=9, column=col).value) for col in range(4, 10))
    total = quantity_sold + all_nozzles

    sheet2['J7'], sheet2['J8'], sheet2['J9'] = quantity_sold, all_nozzles, total
    sheet2['L7'] = total

    # Process diesel calculations
    process_diesel(sheet2, sheet3)
    
    # Process oil calculations
    process_oils(sheet2)
    
    # Process summary section
    process_summary(sheet2)

def process_diesel(sheet2, sheet3):
    """Handle diesel related calculations"""
    # Process Sheet3 row by row for diesel readings
    for row in range(2, sheet3.max_row):
        try:
            # Define column pairs for diesel readings
            column_pairs = [
                ('G', 'C'),  # Machine 1's final goes to next day's initial
                ('H', 'D'),  # Machine 4's final goes to next day's initial
                ('I', 'E')   # Nozzle 3's final goes to next day's initial
            ]
            
            # Copy today's final to tomorrow's initial
            for final_col, initial_col in column_pairs:
                final_value = sheet3.cell(row=row, column=ord(final_col)-ord('A')+1).value
                if final_value is not None:
                    sheet3.cell(row=row+1, column=ord(initial_col)-ord('A')+1).value = final_value
        except Exception as e:
            print(f"Error processing row {row}: {str(e)}")
            continue

    # Process diesel calculations
    for i, col in enumerate(['B', 'C']):
        sheet2[f'{col}15'] = get_number(sheet3[f'{chr(67+i)}2'].value)  # Initial
        sheet2[f'{col}14'] = get_number(sheet3[f'{chr(71+i)}2'].value)  # Final
        sheet2[f'{col}16'] = get_number(sheet2[f'{col}14'].value) - get_number(sheet2[f'{col}15'].value)

    sheet2['D14'] = get_number(sheet2['B16'].value)
    sheet2['D15'] = get_number(sheet2['C16'].value)
    sheet2['D16'] = sheet2['F14'] = get_number(sheet2['D14'].value) + get_number(sheet2['D15'].value)
    sheet2['G14'] = get_number(sheet2['F14'].value) * get_number(sheet2['E27'].value)

def process_oils(sheet2):
    """Handle 2T and W40 oil calculations"""
    # Process 2T oil
    t_rate = get_number(sheet2['E28'].value)
    for row in [21, 22]:
        qty = get_number(sheet2[f'B{row}'].value)
        sheet2[f'C{row}'] = qty * t_rate
    sheet2['C23'] = get_number(sheet2['C21'].value) + get_number(sheet2['C22'].value)

    # Process W40 oil
    w40_rate = get_number(sheet2['E29'].value)
    for row in [21, 22]:
        qty = get_number(sheet2[f'F{row}'].value)
        sheet2[f'G{row}'] = qty * w40_rate
    sheet2['G23'] = get_number(sheet2['G21'].value) + get_number(sheet2['G22'].value)

def process_summary(sheet2):
    """Handle summary section calculations (rows 26-29)"""
    # Column A mappings
    sheet2['A26'] = get_number(sheet2['L7'].value)  # Petrol cash
    sheet2['A27'] = get_number(sheet2['F14'].value)  # Diesel total
    sheet2['A28'] = get_number(sheet2['B21'].value)  # 2T oil cash
    sheet2['A29'] = get_number(sheet2['F21'].value)  # W40 oil cash
    
    # Column B mappings
    sheet2['B26'] = get_number(sheet2['L8'].value)  # Petrol credit
    sheet2['B27'] = get_number(sheet2['G14'].value)  # Diesel amount
    sheet2['B28'] = get_number(sheet2['C21'].value)  # 2T oil amount cash
    sheet2['B29'] = get_number(sheet2['G21'].value)  # W40 oil amount cash

    # Column C mappings (copy from Column B)
    for row in range(26, 30):
        sheet2[f'C{row}'] = get_number(sheet2[f'B{row}'].value)

    # Calculate and copy totals for row 30
    total_b = sum(get_number(sheet2[f'B{row}'].value) for row in range(26, 30))
    sheet2['B30'] = total_b
    sheet2['C30'] = total_b

def update_excel():
    """Main Excel processing function"""
    try:
        wb = load_workbook('Final DB.xlsx')
        sheet1, sheet2, sheet3 = wb['Sheet1'], wb['Sheet2'], wb['Sheet3']
        
        process_sheet2(sheet1, sheet2, sheet3)
        
        wb.save('Final DB - New.xlsx')
        return True
    except Exception as e:
        print(f"Error: {str(e)}")
        return False

@app.route("/", methods=["GET", "POST"])
def upload_file():
    if request.method == "POST":
        if file := request.files.get("file"):
            file.save('Final DB.xlsx')
            success = update_excel()
            return "File processed successfully!" if success else "Error processing file"
    return render_template_string(html_template)

if __name__ == "__main__":
    os.makedirs('templates', exist_ok=True)
    app.run(debug=True)




